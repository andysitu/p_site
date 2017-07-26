from django.shortcuts import render
from django.http import HttpResponse
from django.core import serializers
from .forms import UploadExcelData

from django.utils.encoding import force_text
from django.core.serializers.json import DjangoJSONEncoder

import re

class LazyEncoder(DjangoJSONEncoder):
    def default(self, obj):
        if isinstance(obj, YourCustomType):
            return force_text(obj)
        return super(LazyEncoder, self).default(obj)

def test(request):
    return HttpResponse("HELLO")

def view_map(request):
    test_dic = {'fire': 'ball'}

    json_data = serializers.serialize('json', Test.objects.all(), cls=LazyEncoder)

    return render(
        request,
        'warehouse_map/map.html',
        context={
            'data_dic': json_data,
        }
    )

def upload_excel_data(request):
    if request.method == 'POST':
        upload_excel_data_form = UploadExcelData(request.POST, request.FILES)
        if upload_excel_data_form.is_valid():
            for file in request.FILES.getlist("excel_data_file"):
            # file = request.FILES["excel_data_file"]
                process_excel_file(file)
    else:
        upload_excel_data_form = UploadExcelData()

    return render(
        request,
        'warehouse_map/upload_excel_data.html',
        context={
            "upload_form": upload_excel_data_form,
        }
    )

import openpyxl
from io import BytesIO

def process_excel_file(file):
    filename = file.name
    print(filename)

    file_regex = re.compile('(?P<year>20\d\d)(?P<month>[01]\d)'
                            + '(?P<day>[0123]\d)(?P<hour>[012]\d)'
                            + '(?P<min>\d\d)(?P<sec>\d\d)')
    re_result = re.match(file_regex, filename)
    year = re_result.group("year")
    month = re_result.group("month")
    day = re_result.group("day")
    hour = re_result.group("hour")
    min = re_result.group("min")
    sec = re_result.group("sec")

    print(year, month, day)

    data = file.read()

    wb = openpyxl.load_workbook(filename=BytesIO(data), read_only=True)
    # sheet_list = wb.get_sheet_names()
    sheet = wb.active

    num_rows = sheet.max_row
    num_cols = sheet.max_column

    item_map = {
        "id": "A",
        "description": "AN",
        "sku_name": "AB",
        "customer_code": "AH",
        "quantity": "AQ",
        "ship_quantity": "AC",
        "location": "C",
        "inven_date": "J",
        "rcv": "N",
    }

    for row in range(2, 30):
        value_dic= {}
        row_num = str(row)
        for item_key, col in item_map.items():
            v = sheet[col + row_num].value
            value_dic[item_key] = v
            print(item_key, v)